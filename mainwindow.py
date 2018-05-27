#!/usr/bin/python
# -*- coding: utf-8 -*-

import os
import os.path
import re
import sys
from time import sleep
import urllib.request
from datetime import datetime
import functions

from PyQt5 import uic
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from bs4 import BeautifulSoup
from mysql.connector import Error
from openpyxl import load_workbook
from openpyxl.compat import range

cursor = functions.cursor
conn = functions.conn
Ui_MainWindow, QtBaseClass = uic.loadUiType("mainwindow.ui")


class MyApp(QMainWindow):
    def __init__(self):
        super(MyApp, self).__init__()
        self.ui = Ui_MainWindow()

        self.ui.setupUi(self)
        self.ui.dwnldButton.clicked.connect(self.download)
        self.ui.parseButton.clicked.connect(self.parse_lessons_for_selected_group)
        self.ui.parsAllPushButton.clicked.connect(self.parse_all)
        self.ui.updGlButton.clicked.connect(self.update_group_list)
        self.ui.toTablesButton.clicked.connect(self.to_tables)
        self.ui.titleButton.clicked.connect(self.parse_titles)
        self.ui.tleButton.clicked.connect(self.tle)
        self.ui.tgrButton.clicked.connect(self.tgr)
        self.ui.tpaButton.clicked.connect(self.tpa)
        self.ui.instituteComboBox.activated.connect(self.update_group_list)
        self.ui.weekLabel.setText(str(datetime.now().isocalendar()[1] - 5))
        self.update_institute_list()

    def parse_titles(self):
        # self.ui.centralwidget.setCursor(QCursor(Qt.WaitCursor))

        # print("parsing complete")
        # self.ui.centralwidget.setCursor(QCursor(Qt.ArrowCursor))
        ParseTitlesThread.start()
        ParseTitlesThread.started.connect(self.thread_started)
        ParseTitlesThread.finished.connect(self.thread_finished)

    def update_institute_list(self):
        cursor.execute("SELECT institute FROM paths")
        ins_tuple = cursor.fetchall()
        self.ui.instituteComboBox.clear()
        ins_list = []
        for ins in ins_tuple:
            ins_list.append(ins[0])
        ins_list = sorted(set(ins_list))
        for i in ins_list:
            self.ui.instituteComboBox.addItem(i)

    def update_group_list(self):
        ins = self.ui.instituteComboBox.currentText()
        try:
            cursor.execute("SELECT group_name FROM groups WHERE institute=%s",
                           (ins,))
        except Error as error:
            print(error)

        group_tuple = cursor.fetchall()
        group_list = []
        self.ui.groupComboBox.clear()
        for x in group_tuple:
            group_list.append(x[0])
        group_list = sorted(group_list)
        for group in group_list:
            # strgroup = '-'.join(map(str, group))
            self.ui.groupComboBox.addItem(group)
            # group_list.append(strgroup)
        return group_list

    def download(self):
        DownloadThread.start()
        DownloadThread.started.connect(self.thread_started)
        DownloadThread.finished.connect(self.thread_finished)
        DownloadThread.my_signal.connect(self.my_event)

    def thread_finished(self):
        self.ui.centralwidget.setCursor(QCursor(Qt.ArrowCursor))

    def thread_started(self):
        self.ui.centralwidget.setCursor(QCursor(Qt.WaitCursor))

    def my_event(self):
        print("download complete")

    def to_tables(self):
        group = self.ui.groupComboBox.currentText()
        day = self.ui.daySpinBox.value()
        even = int(self.ui.evenCheckBox.isChecked())
        try:
            cursor.execute("SELECT type, title, teacher, room FROM lessons WHERE day=%s AND even=%s AND `group` = %s",
                           (day, even, group))
            lessons = cursor.fetchall()
        except Error as error:
            print(error)
        # lessons = cursor.fetchall()
        print(lessons)
        for i in range(6):
            for j in range(4):

                if not lessons:
                    lesson = ""
                else:
                    lesson = lessons[i][j]
                if lesson == ("день" or "самостолятельных" or "занятий"):
                    continue
                self.ui.tableWidget1.setItem(i, j, QTableWidgetItem(lesson))
        self.ui.tableWidget1.setColumnWidth(0, 50)
        self.ui.tableWidget1.setColumnWidth(1, 270)
        self.ui.tableWidget1.setColumnWidth(2, 130)
        self.ui.tableWidget1.setColumnWidth(3, 50)

    def parse_lessons(self, groupname):
        self.ui.centralwidget.setCursor(QCursor(Qt.WaitCursor))
        # groupname = self.ui.groupComboBox.currentText()
        print(groupname)
        try:
            cursor.execute("SELECT filename, sheet FROM paths WHERE (groups LIKE %s AND ses='занятия')",
                           # доработать выборку
                           ("%" + groupname + "%",))
        except Error as error:
            print(error)

        fetch = cursor.fetchone();
        fname = fetch[0]
        sheet = fetch[1]
        print(fname, sheet)
        from openpyxl import load_workbook
        wb = load_workbook(filename=fname, read_only=True)
        ws = wb[sheet]

        x = 1
        y = 1
        for row in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=200):
            for cols in row:
                # strvalue = ""
                # strvalue = cols.value
                if groupname in str(cols.value):
                    y = cols.row
                    x = cols.column
                    break

        mir = 4
        mar = mir + 71
        mic = x
        mac = mic + 3
        if not ws.cell(row=y, column=x).value:
            gr = ""
        else:
            gr = ws.cell(row=y, column=x).value
        # gr = ws.cell(row=y, column=x).value
        print(gr)
        number = 1
        for index, row in enumerate(ws.iter_rows(min_row=mir, max_row=mar, min_col=mic, max_col=mac)):
            title = str(row[0].value)
            subgr = 0
            day = index // 12 + 1
            if number > 6:
                number = 1
            if index % 2 == 0:
                even = 0
            else:
                even = 1
            '''if "(1 подгр)" in title:
                print("до: ", title)
                subgr = 1
                title = title.replace('(1 подгр)', '')
                print("после: ", title)
            if "(2 подгр)" in title:
                print("до: ", title)
                subgr = 2
                title = title.replace('(2 подгр)', '')
                print("после: ", title)'''
            title = str(os.linesep.join([s for s in title.splitlines() if s]))
            try:
                cursor.execute("INSERT INTO lessons VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)", (
                    None, gr, day, number, even, title, row[1].value, row[2].value, row[3].value,
                    None, subgr, None))
                conn.commit()
            except Error as error:
                print(error)
            number += even
        self.ui.centralwidget.setCursor(QCursor(Qt.ArrowCursor))

    def parse_all(self):
        group_list = self.update_group_list()
        for group in group_list:
            self.parse_lessons(group)

    def parse_lessons_for_selected_group(self):
        self.parse_lessons(self.ui.groupComboBox.currentText())

    def tle(self):
        self.ui.centralwidget.setCursor(QCursor(Qt.WaitCursor))
        try:
            cursor.execute("TRUNCATE TABLE lessons;")
            conn.commit()
        except Error as error:
            print(error)
        self.ui.centralwidget.setCursor(QCursor(Qt.ArrowCursor))

    def tgr(self):
        self.ui.centralwidget.setCursor(QCursor(Qt.WaitCursor))
        try:
            cursor.execute("TRUNCATE TABLE groups;")
            conn.commit()
        except Error as error:
            print(error)
        self.ui.centralwidget.setCursor(QCursor(Qt.ArrowCursor))

    def tpa(self):
        self.ui.centralwidget.setCursor(QCursor(Qt.WaitCursor))
        try:
            cursor.execute("TRUNCATE TABLE paths;")
            conn.commit()
        except Error as error:
            print(error)
        self.ui.centralwidget.setCursor(QCursor(Qt.ArrowCursor))

    def closeEvent(self, event):
        conn.close()
        print(conn.is_connected())
        event.accept()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    '''app.setStyle('Fusion')
    palette = QPalette()
    palette.setColor(QPalette.Window, QColor(53, 53, 53))
    palette.setColor(QPalette.WindowText, Qt.white)
    palette.setColor(QPalette.Base, QColor(15, 15, 15))
    palette.setColor(QPalette.AlternateBase, QColor(53, 53, 53))
    palette.setColor(QPalette.ToolTipBase, Qt.white)
    palette.setColor(QPalette.ToolTipText, Qt.white)
    palette.setColor(QPalette.Text, Qt.white)
    palette.setColor(QPalette.Button, QColor(53, 53, 53))
    palette.setColor(QPalette.ButtonText, Qt.white)
    palette.setColor(QPalette.BrightText, Qt.red)
    palette.setColor(QPalette.Highlight, QColor(142, 45, 197).lighter())
    palette.setColor(QPalette.HighlightedText, Qt.black)
    app.setPalette(palette)'''
    window = MyApp()
    window.show()
    DownloadThread = functions.DownloadThread()
    ParseTitlesThread = functions.ParseTitlesThread()
    sys.exit(app.exec_())
