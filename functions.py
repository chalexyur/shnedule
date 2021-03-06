import os
import os.path
import re
import urllib.request
import urllib.request
from configparser import ConfigParser
from datetime import datetime

from PyQt5.QtCore import *
from bs4 import BeautifulSoup
from mysql.connector import Error
from mysql.connector import MySQLConnection
from openpyxl import load_workbook
from openpyxl.compat import range


def read_db_config():
    filename = 'config.ini'
    section = 'local-mysql'
    parser = ConfigParser()
    parser.read(filename)
    db = {}
    if parser.has_section(section):
        items = parser.items(section)
        for item in items:
            db[item[0]] = item[1]
    else:
        raise Exception('{0} not found in the {1} file'.format(section, filename))
    return db


dbconfig = read_db_config()
conn = MySQLConnection(**dbconfig)
print(conn.is_connected())
cursor = conn.cursor()
global_groupname: str = "nogroup"

try:
    cursor.execute("""
    create table IF NOT EXISTS lessons
    (
         id       int auto_increment
    primary key,
  `group`  varchar(10)  null,
  day      int(1)       null,
  number   int(1)       null,
  even     tinyint(1)   null,
  title    varchar(999) null,
  type     varchar(20)  null,
  teacher  varchar(99)  null,
  room     varchar(99)  null,
  weeks    varchar(50)  null,
  subgroup int(1)       null,
  campus   varchar(99)  null
);""")
    cursor.execute("""
    create table IF NOT EXISTS paths
    (
       id          int auto_increment
    primary key,
  institute   varchar(50)  null,
  prog        varchar(50)  null,
  course      int          null,
  ses         varchar(50)  null,
  last_update datetime     null,
  past_size   int          null,
  filename    varchar(50)  null,
  sheet       varchar(50)  null,
  title       varchar(999) null,
  university  varchar(50)  null,
  groups      varchar(999) null
);""")
    cursor.execute("""
        create table IF NOT EXISTS groups
        (
          id          int auto_increment
    primary key,
  group_name  varchar(50) not null,
  quantity    int         null,
  institute   varchar(50) null,
  last_update datetime    null,
  constraint groups_group_name_uindex
  unique (group_name)
    );""")
    conn.commit()
except Error as error:
    print(error)


def parse_groups(worksheet, institute):
    ws = worksheet
    groupsstring = ""
    for row in ws.iter_rows(min_row=2, max_row=3, min_col=1, max_col=200):
        for cols in row:
            string = str(cols.value)
            match = re.search(r'\w*[-]\d\d[-]\d\d', string)
            if match:
                string = match[0]
                groupsstring += string + ','
                try:
                    cursor.execute("INSERT IGNORE INTO groups VALUES (%s, %s, %s, %s, %s)",
                                   (None, string, None, institute, None))
                except Error as error:
                    print(error)
    conn.commit()
    return groupsstring


class DownloadThread(QThread):
    my_signal = pyqtSignal()

    def run(self):
        html_page = urllib.request.urlopen('https://www.mirea.ru/education/schedule-main/schedule/').read()
        soup = BeautifulSoup(html_page, "html.parser")
        if not os.path.exists("files/"):
            os.makedirs("files/")

        # for index, link in enumerate(soup.findAll('a', attrs={'href': re.compile(".xls$")})):
        # urllib.request.urlretrieve(link.get('href'), "files/" + str(index) + ".xls")
        for index, link in enumerate(soup.findAll('a', attrs={'href': re.compile(".xlsx$")})):
            print(link.get('href'))
            urllib.request.urlretrieve(link.get('href'), "files/" + str(index) + ".xlsx")
        # for index, link in enumerate(soup.findAll('a', attrs={'href': re.compile(".pdf$")})):
        # urllib.request.urlretrieve(link.get('href'), "files/" + str(index) + ".pdf")
        self.my_signal.emit()


class ParseTitlesThread(QThread):
    def run(self):
        folder = "files/"
        qfiles = len([name for name in os.listdir(folder) if os.path.isfile(os.path.join(folder, name))])
        print(qfiles)
        for i in range(0, 200):
            fpath = folder + str(i) + ".xlsx"
            if not os.path.exists(fpath):
                continue
            print(fpath)
            wb = load_workbook(filename=fpath, read_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=4):
                    for cols in row:
                        value = str(cols.value)
                        if re.match(r"\bр\s*а\s*с\s*п\s*и\s*с\s*а\s*н\s*и\s*е\b", value, re.IGNORECASE):
                            size = os.path.getsize(fpath)
                            course = 0
                            ses = institute = "zero"
                            prog = "бакалавриат/специалитет"
                            university = "МИРЭА"
                            match1 = re.search(r'\w*\d\w*', value)
                            if match1:
                                course = match1[0]
                            match2 = re.search(r'\w*занятий\w*', value)
                            if match2:
                                ses = "занятия"
                            match2 = re.search(r'\w*зачетной\w*', value) or re.search(r'\w*зачетов\w*', value)
                            if match2:
                                ses = "зачётная сессия"
                            match2 = re.search(r'\w*экзаменационной\w*', value)
                            if match2:
                                ses = "экзаменационная сессия"
                            match3 = re.search(r'\w*ИНТЕГУ\w*', value)
                            if match3:
                                institute = "ИНТЕГУ"
                            match3 = re.search(r'\w*КБиСП\w*', value) or re.search(r'\w*КБСП\w*', value)
                            if match3:
                                institute = "КБиСП"
                            match3 = re.search(r'\w*кибернетики\w*', value)
                            if match3:
                                institute = "ИК"
                            match3 = re.search(r'\w*\bФизико\s*-\s*технологического\w*\b', value) or re.search(
                                r'\w*ФТИ\w*', value)
                            if match3:
                                institute = "ФТИ"
                            match3 = re.search(r'\w*\bИТ\s*\w*\b', value)
                            if match3:
                                institute = "ИТ"
                            match3 = re.search(r'\w*РТС\w*', value)
                            if match3:
                                institute = "РТС"
                            match3 = re.search(r'\w*ИЭС\w*', value)
                            if match3:
                                institute = "ИЭС"
                            match3 = re.search(r'\w*ИЭП\w*', value)
                            if match3:
                                institute = "ИЭП"
                            match3 = re.search(r'\w*ВЗО\w*', value)
                            if match3:
                                institute = "ИВЗО"
                            match3 = re.search(r'\w*ИУСТРО\w*', value)
                            if match3:
                                institute = "ИУСТРО"

                            match4 = re.search(r'\w*магистратуры\w*', value)
                            if match4:
                                prog = "магистратура"
                            groupsstring = parse_groups(ws, institute)
                            cursor.execute("INSERT INTO paths VALUES (%s,%s, %s, %s, %s, %s, %s ,%s,%s,%s,%s,%s)",
                                           (
                                               None, institute, prog, course, ses, datetime.now(), size, fpath, sheet,
                                               None,
                                               university, groupsstring))
                            conn.commit()


class ParseLessonsThread(QThread):
    def run(self):
        groupname = global_groupname
        print(groupname)
        try:
            cursor.execute("SELECT filename, sheet FROM paths WHERE (groups LIKE %s AND ses='занятия')",
                           ("%" + groupname + "%",))
        except Error as error:
            print(error)

        fetch = cursor.fetchone();
        fname = fetch[0]
        sheet = fetch[1]
        print(fname, sheet)
        from openpyxl import load_workbook
        wb = load_workbook(filename=fname, read_only=True)
        ws = wb[sheet]

        x = 1
        y = 1
        for row in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=200):
            for cols in row:
                if groupname in str(cols.value):
                    y = cols.row
                    x = cols.column
                    break

        mir = 4
        mar = mir + 71
        mic = x
        mac = mic + 3
        if not ws.cell(row=y, column=x).value:
            gr = ""
        else:
            gr = ws.cell(row=y, column=x).value
        print(gr)
        number = 1
        for index, row in enumerate(ws.iter_rows(min_row=mir, max_row=mar, min_col=mic, max_col=mac)):
            title = str(row[0].value)
            subgr = 0
            day = index // 12 + 1
            if number > 6:
                number = 1
            if index % 2 == 0:
                even = 0
            else:
                even = 1
            title = str(os.linesep.join([s for s in title.splitlines() if s]))
            try:
                cursor.execute("INSERT INTO lessons VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)", (
                    None, gr, day, number, even, title, row[1].value, row[2].value, row[3].value,
                    None, subgr, None))
                conn.commit()
            except Error as error:
                print(error)
            number += even
