#!/usr/bin/python
# -*- coding: utf-8 -*-

import os
import os.path
import re
import sys
from time import sleep
import urllib.request
from datetime import datetime
import functions

from PyQt5 import uic
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from bs4 import BeautifulSoup
from mysql.connector import Error
from mysql.connector import MySQLConnection
from openpyxl import load_workbook
from openpyxl.compat import range

dbconfig = functions.read_db_config()
conn = MySQLConnection(**dbconfig)
print(conn.is_connected())
cursor = conn.cursor()

try:

    cursor.execute("""
    create table IF NOT EXISTS lessons
    (
         id       int auto_increment
    primary key,
  `group`  varchar(10)  null,
  day      int(1)       null,
  number   int(1)       null,
  even     tinyint(1)   null,
  title    varchar(999) null,
  type     varchar(20)  null,
  teacher  varchar(99)  null,
  room     varchar(99)  null,
  weeks    varchar(50)  null,
  subgroup int(1)       null,
  campus   varchar(99)  null
);""")
    cursor.execute("""
    create table IF NOT EXISTS paths
    (
       id          int auto_increment
    primary key,
  institute   varchar(50)  null,
  prog        varchar(50)  null,
  course      int          null,
  ses         varchar(50)  null,
  last_update datetime     null,
  past_size   int          null,
  filename    varchar(50)  null,
  sheet       varchar(50)  null,
  title       varchar(999) null,
  university  varchar(50)  null,
  groups      varchar(999) null
);""")
    cursor.execute("""
        create table IF NOT EXISTS groups
        (
          id          int auto_increment
    primary key,
  group_name  varchar(50) not null,
  quantity    int         null,
  institute   varchar(50) null,
  last_update datetime    null,
  constraint groups_group_name_uindex
  unique (group_name)
    );""")
    conn.commit()
except Error as error:
    print(error)


def parse_groups(worksheet, institute):
    ws = worksheet
    groupsstring = ""
    for row in ws.iter_rows(min_row=2, max_row=3, min_col=1, max_col=200):
        for cols in row:
            string = str(cols.value)
            match = re.search(r'\w*[-]\d\d[-]\d\d', string)
            if match:
                # print(string)
                string = match[0]
                # print(gr_code)
                # print(string)
                groupsstring += string + ','
                try:
                    # cursor.execute("INSERT INTO groups VALUES (%s, %s, %s, %s, %s, %s, %s,%s)",
                    # (None, group[0], group[1], int(group[2]), None, None, None, None))
                    cursor.execute("INSERT IGNORE INTO groups VALUES (%s, %s, %s, %s, %s)",
                                   (None, string, None, institute, None))
                    # (group[0], group[1], group[2]))
                    # cursor.execute("REPLACE INTO groups SET name=%s, code=%s, year=%s", (group[0], group[1], group[2]))
                    # cursor.execute("INSERT INTO groups SET name=%s", (group[0]))
                except Error as error:
                    print(error)
    conn.commit()
    return groupsstring


Ui_MainWindow, QtBaseClass = uic.loadUiType("mainwindow.ui")


class ExecuteThread(QThread):
    my_signal = pyqtSignal()

    def run(self):
        html_page = urllib.request.urlopen('https://www.mirea.ru/education/schedule-main/schedule/').read()
        soup = BeautifulSoup(html_page, "html.parser")
        if not os.path.exists("files/"):
            os.makedirs("files/")

        # for index, link in enumerate(soup.findAll('a', attrs={'href': re.compile(".xls$")})):
        # urllib.request.urlretrieve(link.get('href'), "files/" + str(index) + ".xls")
        for index, link in enumerate(soup.findAll('a', attrs={'href': re.compile(".xlsx$")})):
            print(link.get('href'))
            urllib.request.urlretrieve(link.get('href'), "files/" + str(index) + ".xlsx")
            sleep(2)
        # for index, link in enumerate(soup.findAll('a', attrs={'href': re.compile(".pdf$")})):
        # urllib.request.urlretrieve(link.get('href'), "files/" + str(index) + ".pdf")
        self.my_signal.emit()
        pass


def movetothread():
    print(1)
    runnable = Runnable()
    QThreadPool.globalInstance().start(runnable)


class MyApp(QMainWindow):
    def __init__(self):
        # QThread.__init__(self)
        super(MyApp, self).__init__()
        self.ui = Ui_MainWindow()

        self.ui.setupUi(self)
        self.ui.dwnldButton.clicked.connect(self.download)
        self.ui.parseButton.clicked.connect(self.parse_lessons_for_selected_group)
        self.ui.parsAllPushButton.clicked.connect(self.parse_all)
        self.ui.updGlButton.clicked.connect(self.update_group_list)
        self.ui.toTablesButton.clicked.connect(self.to_tables)
        self.ui.titleButton.clicked.connect(self.parse_titles)
        self.ui.tleButton.clicked.connect(self.tle)
        self.ui.tgrButton.clicked.connect(self.tgr)
        self.ui.tpaButton.clicked.connect(self.tpa)
        self.ui.instituteComboBox.activated.connect(self.update_group_list)
        self.ui.weekLabel.setText(str(datetime.now().isocalendar()[1] - 5))
        self.update_institute_list()

    def parse_titles(self):
        self.ui.centralwidget.setCursor(QCursor(Qt.WaitCursor))
        folder = "files/"
        qfiles = len([name for name in os.listdir(folder) if os.path.isfile(os.path.join(folder, name))])
        print(qfiles)
        for i in range(0, 99):
            fpath = folder + str(i) + ".xlsx"
            if not os.path.exists(fpath):
                continue
            print(fpath)
            wb = load_workbook(filename=fpath, read_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=4):
                    for cols in row:
                        value = str(cols.value)
                        if re.match(r"\bр\s*а\s*с\s*п\s*и\s*с\s*а\s*н\s*и\s*е\b", value, re.IGNORECASE):
                            size = os.path.getsize(fpath)
                            course = 0
                            ses = institute = "zero"
                            prog = "бакалавриат/специалитет"
                            university = "МИРЭА"
                            match1 = re.search(r'\w*\d\w*', value)
                            if match1:
                                course = match1[0]
                            match2 = re.search(r'\w*занятий\w*', value)
                            if match2:
                                ses = "занятия"
                            match2 = re.search(r'\w*зачетной\w*', value) or re.search(r'\w*зачетов\w*', value)
                            if match2:
                                ses = "зачётная сессия"
                            match2 = re.search(r'\w*экзаменационной\w*', value)
                            if match2:
                                ses = "экзаменационная сессия"
                            match3 = re.search(r'\w*ИНТЕГУ\w*', value)
                            if match3:
                                institute = "ИНТЕГУ"
                            match3 = re.search(r'\w*КБиСП\w*', value)
                            if match3:
                                institute = "КБиСП"
                            match3 = re.search(r'\w*кибернетики\w*', value)
                            if match3:
                                institute = "ИК"
                            match3 = re.search(r'\w*\bФизико\s*-\s*технологического\w*\b', value) or re.search(r'\w*ФТИ\w*', value)
                            if match3:
                                institute = "ФТИ"
                            match3 = re.search(r'\w*\bИТ\s*\w*\b', value)
                            if match3:
                                institute = "ИТ"
                            match3 = re.search(r'\w*РТС\w*', value)
                            if match3:
                                institute = "РТС"
                            match3 = re.search(r'\w*ИЭС\w*', value)
                            if match3:
                                institute = "ИЭС"
                            match3 = re.search(r'\w*ИЭП\w*', value)
                            if match3:
                                institute = "ИЭП"
                            match3 = re.search(r'\w*ВЗО\w*', value)
                            if match3:
                                institute = "ИВЗО"
                            match3 = re.search(r'\w*ИУСТРО\w*', value)
                            if match3:
                                institute = "ИУСТРО"

                            match4 = re.search(r'\w*магистратуры\w*', value)
                            if match4:
                                prog = "магистратура"
                            groupsstring = parse_groups(ws, institute)
                            cursor.execute("INSERT INTO paths VALUES (%s,%s, %s, %s, %s, %s, %s ,%s,%s,%s,%s,%s)",
                                           (
                                               None, institute, prog, course, ses, datetime.now(), size, fpath, sheet,
                                               None,
                                               university, groupsstring))
                            conn.commit()
        self.ui.centralwidget.setCursor(QCursor(Qt.ArrowCursor))

    def update_institute_list(self):
        cursor.execute("SELECT institute FROM paths")
        ins_tuple = cursor.fetchall()
        self.ui.instituteComboBox.clear()
        ins_list = []
        for ins in ins_tuple:
            ins_list.append(ins[0])
        ins_list = sorted(set(ins_list))
        for i in ins_list:
            self.ui.instituteComboBox.addItem(i)

    def update_group_list(self):
        ins = self.ui.instituteComboBox.currentText()
        try:
            cursor.execute("SELECT group_name FROM groups WHERE institute=%s",
                           (ins,))
        except Error as error:
            print(error)

        group_tuple = cursor.fetchall()
        group_list = []
        self.ui.groupComboBox.clear()
        for x in group_tuple:
            group_list.append(x[0])
        group_list = sorted(group_list)
        for group in group_list:
            # strgroup = '-'.join(map(str, group))
            self.ui.groupComboBox.addItem(group)
            # group_list.append(strgroup)
        return group_list

    def download(self):

        # Thread(target=self.downloading_thread()).start

        test.start()
        test.started.connect(self.thread_started)
        test.finished.connect(self.thread_finished)
        test.my_signal.connect(self.my_event)

    def thread_finished(self):
        self.ui.centralwidget.setCursor(QCursor(Qt.ArrowCursor))

    def thread_started(self):
        self.ui.centralwidget.setCursor(QCursor(Qt.WaitCursor))

    def my_event(self):
        print("sapsan")
        pass

    def download_thread(self):
        html_page = urllib.request.urlopen('https://www.mirea.ru/education/schedule-main/schedule/').read()
        soup = BeautifulSoup(html_page, "html.parser")
        if not os.path.exists("files/"):
            os.makedirs("files/")

        # for index, link in enumerate(soup.findAll('a', attrs={'href': re.compile(".xls$")})):
        # urllib.request.urlretrieve(link.get('href'), "files/" + str(index) + ".xls")
        for index, link in enumerate(soup.findAll('a', attrs={'href': re.compile(".xlsx$")})):
            print(link.get('href'))
            urllib.request.urlretrieve(link.get('href'), "files/" + str(index) + ".xlsx")
            sleep(2)
        # for index, link in enumerate(soup.findAll('a', attrs={'href': re.compile(".pdf$")})):
        # urllib.request.urlretrieve(link.get('href'), "files/" + str(index) + ".pdf")

    def to_tables(self):
        group = self.ui.groupComboBox.currentText()
        day = self.ui.daySpinBox.value()
        even = int(self.ui.evenCheckBox.isChecked())
        try:
            cursor.execute("SELECT type, title, teacher, room FROM lessons WHERE day=%s AND even=%s AND `group` = %s",
                           (day, even, group))
            lessons = cursor.fetchall()
        except Error as error:
            print(error)
        # lessons = cursor.fetchall()
        print(lessons)
        for i in range(6):
            for j in range(4):

                if not lessons:
                    lesson = ""
                else:
                    lesson = lessons[i][j]
                if lesson == ("день" or "самостолятельных" or "занятий"):
                    continue
                self.ui.tableWidget1.setItem(i, j, QTableWidgetItem(lesson))
        self.ui.tableWidget1.setColumnWidth(0, 50)
        self.ui.tableWidget1.setColumnWidth(1, 270)
        self.ui.tableWidget1.setColumnWidth(2, 130)
        self.ui.tableWidget1.setColumnWidth(3, 50)

    def parse_lessons(self, groupname):
        self.ui.centralwidget.setCursor(QCursor(Qt.WaitCursor))
        # groupname = self.ui.groupComboBox.currentText()
        print(groupname)
        try:
            cursor.execute("SELECT filename, sheet FROM paths WHERE (groups LIKE %s AND ses='занятия')",
                           # доработать выборку
                           ("%" + groupname + "%",))
        except Error as error:
            print(error)

        fetch = cursor.fetchone();
        fname = fetch[0]
        sheet = fetch[1]
        print(fname, sheet)
        from openpyxl import load_workbook
        wb = load_workbook(filename=fname, read_only=True)
        ws = wb[sheet]

        x = 1
        y = 1
        for row in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=200):
            for cols in row:
                # strvalue = ""
                # strvalue = cols.value
                if groupname in str(cols.value):
                    y = cols.row
                    x = cols.column
                    break

        mir = 4
        mar = mir + 71
        mic = x
        mac = mic + 3
        if not ws.cell(row=y, column=x).value:
            gr = ""
        else:
            gr = ws.cell(row=y, column=x).value
        # gr = ws.cell(row=y, column=x).value
        print(gr)
        number = 1
        for index, row in enumerate(ws.iter_rows(min_row=mir, max_row=mar, min_col=mic, max_col=mac)):
            title = str(row[0].value)
            subgr = 0
            day = index // 12 + 1
            if number > 6:
                number = 1
            if index % 2 == 0:
                even = 0
            else:
                even = 1
            '''if "(1 подгр)" in title:
                print("до: ", title)
                subgr = 1
                title = title.replace('(1 подгр)', '')
                print("после: ", title)
            if "(2 подгр)" in title:
                print("до: ", title)
                subgr = 2
                title = title.replace('(2 подгр)', '')
                print("после: ", title)'''
            title = str(os.linesep.join([s for s in title.splitlines() if s]))
            try:
                cursor.execute("INSERT INTO lessons VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)", (
                    None, gr, day, number, even, title, row[1].value, row[2].value, row[3].value,
                    None, subgr, None))
                conn.commit()
            except Error as error:
                print(error)
            number += even
        self.ui.centralwidget.setCursor(QCursor(Qt.ArrowCursor))

    def parse_all(self):
        group_list = self.update_group_list()
        for group in group_list:
            self.parse_lessons(group)

    def parse_lessons_for_selected_group(self):
        self.parse_lessons(self.ui.groupComboBox.currentText())

    def tle(self):
        self.ui.centralwidget.setCursor(QCursor(Qt.WaitCursor))
        try:
            cursor.execute("TRUNCATE TABLE lessons;")
            conn.commit()
        except Error as error:
            print(error)
        self.ui.centralwidget.setCursor(QCursor(Qt.ArrowCursor))

    def tgr(self):
        self.ui.centralwidget.setCursor(QCursor(Qt.WaitCursor))
        try:
            cursor.execute("TRUNCATE TABLE groups;")
            conn.commit()
        except Error as error:
            print(error)
        self.ui.centralwidget.setCursor(QCursor(Qt.ArrowCursor))

    def tpa(self):
        self.ui.centralwidget.setCursor(QCursor(Qt.WaitCursor))
        try:
            cursor.execute("TRUNCATE TABLE paths;")
            conn.commit()
        except Error as error:
            print(error)
        self.ui.centralwidget.setCursor(QCursor(Qt.ArrowCursor))

    def closeEvent(self, event):
        conn.close()
        print(conn.is_connected())
        event.accept()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    '''app.setStyle('Fusion')
    palette = QtGui.QPalette()
    palette.setColor(QtGui.QPalette.Window, QtGui.QColor(53, 53, 53))
    palette.setColor(QtGui.QPalette.WindowText, QtCore.Qt.white)
    palette.setColor(QtGui.QPalette.Base, QtGui.QColor(15, 15, 15))
    palette.setColor(QtGui.QPalette.AlternateBase, QtGui.QColor(53, 53, 53))
    palette.setColor(QtGui.QPalette.ToolTipBase, QtCore.Qt.white)
    palette.setColor(QtGui.QPalette.ToolTipText, QtCore.Qt.white)
    palette.setColor(QtGui.QPalette.Text, QtCore.Qt.white)
    palette.setColor(QtGui.QPalette.Button, QtGui.QColor(53, 53, 53))
    palette.setColor(QtGui.QPalette.ButtonText, QtCore.Qt.white)
    palette.setColor(QtGui.QPalette.BrightText, QtCore.Qt.red)
    palette.setColor(QtGui.QPalette.Highlight, QtGui.QColor(142, 45, 197).lighter())
    palette.setColor(QtGui.QPalette.HighlightedText, QtCore.Qt.black)
    app.setPalette(palette)'''
    window = MyApp()
    window.show()
    test = ExecuteThread()
    sys.exit(app.exec_())
